import datetime
import httplib2
import os
import json
import pprint

from apiclient import discovery
import oauth2client
from oauth2client import client
from oauth2client import tools

from datetime import datetime
from datetime import date
from datetime import timedelta
import time

# try:
#     import argparse
#     flags = argparse.ArgumentParser(parents=[tools.argparser]).parse_args()
# except ImportError:
#     flags = None

# This is HelloCal calendar for fzubair@gmail.com, Change it as appropriate.
SCOPES = 'https://www.googleapis.com/auth/calendar'
CLIENT_SECRET_FILE = ''
APPLICATION_NAME = ''
CAL_ID = ''

def get_credentials(args):
    """Gets valid user credentials from storage.

    If nothing has been stored, or if the stored credentials are invalid,
    the OAuth2 flow is completed to obtain the new credentials.

    Returns:
        Credentials, the obtained credential.
    """
    home_dir = os.path.expanduser('~')
    credential_dir = os.path.join(home_dir, '.credentials')
    if not os.path.exists(credential_dir):
        os.makedirs(credential_dir)
    credential_path = os.path.join(credential_dir,
                                   'calendar-quickstart.json')
    store = oauth2client.file.Storage(credential_path)

    credentials = store.get()
    if not credentials or credentials.invalid:
        print(CLIENT_SECRET_FILE)
        flow = client.flow_from_clientsecrets(CLIENT_SECRET_FILE, SCOPES)
        flow.user_agent = APPLICATION_NAME
        if args:
            credentials = tools.run_flow(flow, store, args)
        else: # Needed only for compatibility with Python 2.6
            credentials = tools.run(flow, store)
        print('Storing credentials to ' + credential_path)
    return credentials


days_of_week = ["M", "T", "W", "R", "F", "S", "Sun"]
prof_list = ["FORTIER, RANDY", "BRADBURY, JEREMY", "SZLICHTA, JAROSLAW", "QURESHI, FAISAL", "PU, KEN", "GREEN, MARK", "COLLINS, CHRISTOPHER"]

def find_date(start_date, d):
    offset = start_date.weekday()

    i = 0
    for x in days_of_week:
        if x == d:
            break
        i = i + 1

    i = (i - offset) % 7
    return start_date + timedelta(days=i)


# def get_event_list(service, calendarId):
#     now = datetime.utcnow().isoformat() + 'Z'
#     event_result = service.events().list(calendarId=calendarId, timeMin=now, orderBy='startTime', singleEvents=True).execute()
#     events = event_result.get('items', [])
#     if not events:
#         print 'No upcoming events found'
#     else:
#         for event in events:
#             # start = event['start'].get('dateTime', event['start'].get('date'))
#             # print start, event['summary']
#             print event['summary'], event['id'], event['start']
#             print "\n\n"


def create_cal_event(sdate, edate, course, name, day, btime, etime, loc):

    d = find_date(sdate, day)
    ds = datetime.combine(d, datetime.strptime(btime, "%H%M").time())
    de = datetime.combine(d, datetime.strptime(etime, "%H%M").time())
    # n = datetime.now()
    # u = datetime.utcnow()
    # utcoffset = n - u
    # ds = ds - utcoffset
    # de = de - utcoffset

    # print ds
    # print de
    # print n
    # print u
    # print utcoffset
    ds = ds.isoformat()
    de = de.isoformat()
    # print ds[:19]
    # print de[:19]
    ed = edate.isoformat()

    return {
        'kind': 'calendar#event',
        'summary': '%s %s' % (course, name),
        'loc': loc,
        'end': {'dateTime': de[:19], 'timeZone': 'America/Toronto'},
        'start': {'dateTime': ds[:19], 'timeZone': 'America/Toronto'},
        "recurrence": [
            "RRULE:FREQ=WEEKLY;UNTIL=%s%s%sT235959Z" % (ed[0:4],ed[5:7],ed[8:10]),
        ],
    }

def main():
    # Lets process commandline arguments
    import argparse
    # parser = argparse.ArgumentParser()
    parser = argparse.ArgumentParser(parents=[tools.argparser])
    parser.add_argument("--no-gcal",action="store_true",default=False,help="Lists relevant events found in the banner file.  Doesn't add anything to the Google calendar.")
    parser.add_argument("--add",action="store_true",default=False,help="Add all events into the calendar")
    parser.add_argument("--delete-all",action="store_true",default=False,help="Delete all events stored in the calendar")
    parser.add_argument("-l","--list",action="store_true",default=False,help="Lists events stored in the calendar")
    parser.add_argument("-d","--debug", action="store_true",default=False, help="Sets debug flag.")
    parser.add_argument("bannerfile", help="Name of the banner excel file.")
    args = parser.parse_args()

    if args.debug:
        print(args)

    # parser = argparse.ArgumentParser(parents=[tools.argparser])
    # flags = parser.parse_args()

    # Lets connect with gcal
    if not args.no_gcal:
        credentials = get_credentials(args)
        http = credentials.authorize(httplib2.Http())
        service = discovery.build('calendar', 'v3', http=http)

    # get_event_list(service, CAL_ID)
    # exit()

    # # Code to get all possible calendars
    # page_token = None
    # while True:
    #     calendar_list = service.calendarList().list(pageToken=page_token).execute()
    #     for calendar_list_entry in calendar_list['items']:
    #         print calendar_list_entry['summary'], calendar_list_entry['id']
    #     page_token = calendar_list.get('nextPageToken')
    #     if not page_token:
    #         break
    # exit()

    # We simply list the events stored int he calendar
    if args.list:
        if args.no_gcal:
            print("Cannot connect to Google Calendar with --no-gcal option")
            parser.print_help()
            exit()

        now = datetime.utcnow().isoformat() + 'Z'
        event_result = service.events().list(calendarId=CAL_ID, timeMin=now, singleEvents=True).execute()
        events = event_result.get('items', [])
        if not events:
            print('No upcoming events found')
        else:
            for event in events:
                start = event['start'].get('dateTime', event['start'].get('date'))
                print(start, event['summary'])
        exit()

    # We want to delete all events
    if args.delete_all:
        if args.no_gcal:
            print("Cannot connect to Google Calendar with --no-gcal option")
            parser.print_help()
            exit()

        now = datetime.utcnow().isoformat() + 'Z'
        event_result = service.events().list(calendarId=CAL_ID, timeMin=now, singleEvents=True).execute()
        events = event_result.get('items', [])
        if not events:
            print('No events found.  Nothing to delete')
        else:
            for event in events:
                start = event['start'].get('dateTime', event['start'].get('date'))
                print("Deleting", start, event['summary'], event['id'])
                service.events().delete(calendarId=CAL_ID, eventId=event['id']).execute()
                time.sleep(1)
        exit()

    # start_date = date(2015, 9, 10)
    # if True:
    #     for d in days_of_week:
    #         print d, find_date(start_date, d)

    # Load up banner file
    from openpyxl import load_workbook
    bannerfile = args.bannerfile
    wb = load_workbook(bannerfile, keep_vba=True)
    try:
        ws = wb.get_sheet_by_name("CSCI") # First attempt to find the CSCI sheet, this is generated by my VBA script
    except:
        ws = 0
        print('Worksheet CSCI not found in the xls file.')

    if not ws:
        try: 
            ws = wb.get_sheet_by_name("ALL") # If CSCI not found, go for ALL
        except:
            print('Worksheet ALL not found in xls file.')
            exit()

    c_idx = {} # Lets try to find the column headings
    row = 1
    col = 1
    v = ws.cell(row=row, column=col).value
    while not v == None:
        c_idx[v] = col
        col = col + 1
        v = ws.cell(row = row, column=col).value
    if args.debug:
        print(c_idx)

    row = 2 # Now lets iterate over the rows to get the items that we want
    while not ws.cell(row=row,column=1).value == None:
        if not ws.cell(row=row, column=c_idx['Subject']).value == 'CSCI':
            row = row + 1
            continue

        if args.debug:
            print('A CSCI course found')

        name = ws.cell(row=row, column=c_idx['Professor']).value
        if not name:
            continue

        print(name)

        if name.upper() in prof_list and ws.cell(row=row, column=c_idx['Sch']).value == 'LEC':
            course = ws.cell(row=row, column=c_idx['Course']).value
            day = ws.cell(row=row, column=c_idx['Day']).value
            btime = ws.cell(row=row, column=c_idx['BTime']).value
            etime = ws.cell(row=row, column=c_idx['ETime']).value
            loc = ws.cell(row=row, column=c_idx['Room']).value
            sdate = ws.cell(row=row, column=c_idx['Start Date']).value
            edate = ws.cell(row=row, column=c_idx['End Date']).value

            if not args.add:
                print("Found", course, name, day, btime, etime, loc, sdate, edate)

            event = create_cal_event(sdate, edate, course, name, day, btime, etime, loc)
            if args.debug:
                print("Event json", event)

            if args.add:
                if args.no_gcal:
                    print("Cannot connect to Google Calendar with --no-gcal option")
                    parser.print_help()
                    exit()

                print("Adding", event)
                httplib2.debuglevel = 4
                service.events().insert(calendarId=CAL_ID, body=event).execute()
                time.sleep(.25)

        row = row + 1


if __name__ == "__main__":
    main()
